"""Export helpers — DataFrame or SQL cursor → FastAPI StreamingResponse.

Three formats (xlsx, csv, zip-of-csv) and two entry points each:

* `dataframe_to_{xlsx,csv,zip}_response` — for small/medium exports where
  loading the whole result into a DataFrame is fine.
* `cursor_to_{xlsx,csv,zip}_response` — for large exports (e.g. signals at
  12 months ~ 379K rows). Streams rows straight from a SQLAlchemy result
  without materializing the whole DataFrame in RAM. XLSX writes via
  openpyxl write_only; CSV and ZIP write via Python's csv module.

Use `dispatch_export_response(fmt, ...)` to pick a format at request time.
"""
from __future__ import annotations

import csv as _csv
import io
import os
import re
import tempfile
import zipfile

import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
ZIP_MEDIA_TYPE = "application/zip"

# Formats accepted by the format-aware export endpoints.
SUPPORTED_FORMATS = ("xlsx", "csv", "zip")


def _cell_value(v):
    """Coerce pandas-specific sentinels (NaN, NaT, pd.NA) and Timestamps for openpyxl."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    return v


def _new_workbook(columns, sheet_name, column_widths: dict | None = None):
    """Create a write-only workbook with header row + frozen panes.

    `column_widths` optionally maps a column name to a width override (in Excel
    character units). Unset columns default to max(12, len(header)+2).

    NOTE: openpyxl write-only mode requires column dimensions to be set
    BEFORE the first ws.append() call — otherwise they silently reset to the
    default. A previous version of this code set them after the header row
    and got default widths everywhere.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name[:31])

    # Set column widths FIRST, before any append.
    widths = column_widths or {}
    for col_idx, col_name in enumerate(columns, start=1):
        w = widths.get(col_name)
        if w is None:
            w = max(12, len(str(col_name)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Now the header row + freeze.
    bold = Font(bold=True)
    header = []
    for col_name in columns:
        c = WriteOnlyCell(ws, value=col_name)
        c.font = bold
        header.append(c)
    ws.append(header)
    ws.freeze_panes = "A2"
    return wb, ws


def _autosize_width(header: str, sample_values, fmt: str | None) -> float:
    """Estimate column width for auto-size mode.

    Width = max char-length across the header and every sample value's formatted
    display, padded by 2 for header bold + 1 for safety. Bounded to [10, 60].
    Format strings that add characters (percent, signed-percent, parens) are
    accounted for via a small fudge factor.
    """
    def _fmt_len(v) -> int:
        if v is None:
            return 0
        if isinstance(v, float):
            if v != v:  # NaN
                return 0
            # heuristic: assume 2-decimal printing covers most format strings
            s = f"{v:,.2f}"
        else:
            s = str(v)
        return len(s)

    longest_value = max((_fmt_len(v) for v in sample_values), default=0)
    fudge = 0
    if fmt:
        if "%" in fmt:
            fudge += 2   # % sign
        if "+" in fmt or "(" in fmt:
            fudge += 2   # sign / parens
        if "Red" in fmt:
            fudge += 1
    width = max(len(str(header)) + 2, longest_value + fudge + 2)
    return min(60.0, max(10.0, float(width)))


def _save_and_stream(wb, filename: str) -> StreamingResponse:
    """Save the workbook to a temp file and stream it back; delete after read.

    Saving to a file (not BytesIO) keeps peak RAM low for large workbooks.
    """
    tmp = tempfile.NamedTemporaryFile(prefix="vesign_export_", suffix=".xlsx", delete=False)
    tmp.close()
    try:
        wb.save(tmp.name)
    except Exception:
        os.unlink(tmp.name)
        raise

    def _file_iter():
        try:
            with open(tmp.name, "rb") as f:
                while True:
                    chunk = f.read(65536)
                    if not chunk:
                        break
                    yield chunk
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

    safe_name = re.sub(r'[\r\n"\\]', '', filename)
    headers = {
        "content-disposition": f'attachment; filename="{safe_name}.xlsx"',
    }
    return StreamingResponse(_file_iter(), media_type=XLSX_MEDIA_TYPE, headers=headers)


def dataframe_to_xlsx_response(
    df: pd.DataFrame,
    filename: str,
    sheet_name: str = "Sheet1",
    column_formats: dict | None = None,
) -> StreamingResponse:
    """Build an XLSX from `df` and return it as a download attachment.

    `filename` should NOT include the .xlsx extension — it's added here.
    Empty DataFrames produce a header-only workbook (caller's intent: "no rows
    matched my filters" should not be a hard error).

    `column_formats` optionally maps a column name to an Excel number-format
    string applied to every data cell in that column (e.g. {"sell_date":
    "dd/mm/yy", "return_pct": "0.00%"}).
    """
    columns = list(df.columns)
    wb, ws = _new_workbook(columns, sheet_name)

    formats = column_formats or {}
    col_fmts = [formats.get(name) for name in columns]

    for row in df.itertuples(index=False, name=None):
        cells = []
        for fmt, raw in zip(col_fmts, row):
            cell = WriteOnlyCell(ws, value=_cell_value(raw))
            if fmt:
                cell.number_format = fmt
            cells.append(cell)
        ws.append(cells)

    return _save_and_stream(wb, filename)


def cursor_to_xlsx_response(
    conn,
    sql,
    params: dict,
    filename: str,
    sheet_name: str = "Sheet1",
    column_formats: dict | None = None,
    date_columns: tuple = (),
    auto_size: bool = False,
) -> StreamingResponse:
    """Stream SQL rows directly into an XLSX, bypassing pandas.

    Used by large exports where loading every row into a DataFrame would OOM.
    `conn` is an active SQLAlchemy connection. `sql` is a `text()` clause.

    `column_formats` maps column name → Excel number-format string applied to
    every data cell in that column. `date_columns` is a list of column names
    whose string values (`YYYY-MM-DD` or `YYYY-MM-DD HH:MM:SS...`) should be
    parsed into Python `date` objects before write so Excel renders them as
    real dates.

    When `auto_size=True`, the cursor is fully materialized into memory first
    so per-column widths can be computed from actual data values. Safe for
    exports with a hard row cap (e.g. 200k); avoid for unbounded queries.
    """
    from datetime import date, datetime

    result = conn.execute(sql, params)
    columns = list(result.keys())

    formats = column_formats or {}
    col_fmts = [formats.get(name) for name in columns]
    is_date = [name in set(date_columns) for name in columns]

    def _parse_date(v):
        if v is None or isinstance(v, (date, datetime)):
            return v
        s = str(v)[:10]
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return v

    # Auto-size path: materialize all rows, compute widths, then write.
    if auto_size:
        rows = list(result)
        column_widths = {}
        for i, name in enumerate(columns):
            samples = [r[i] for r in rows]
            column_widths[name] = _autosize_width(name, samples, col_fmts[i])
        wb, ws = _new_workbook(columns, sheet_name, column_widths=column_widths)

        has_formatting = any(col_fmts) or any(is_date)
        for row in rows:
            if not has_formatting:
                ws.append([_cell_value(v) for v in row])
                continue
            cells = []
            for fmt, dt, raw in zip(col_fmts, is_date, row):
                value = _parse_date(raw) if dt else _cell_value(raw)
                if fmt:
                    cell = WriteOnlyCell(ws, value=value)
                    cell.number_format = fmt
                    cells.append(cell)
                else:
                    cells.append(value)
            ws.append(cells)
        return _save_and_stream(wb, filename)

    # Streaming path (default): no auto-size, write rows as they arrive.
    wb, ws = _new_workbook(columns, sheet_name)
    has_formatting = any(col_fmts) or any(is_date)

    for row in result:
        if not has_formatting:
            ws.append([_cell_value(v) for v in row])
            continue
        cells = []
        for fmt, dt, raw in zip(col_fmts, is_date, row):
            value = _parse_date(raw) if dt else _cell_value(raw)
            if fmt:
                cell = WriteOnlyCell(ws, value=value)
                cell.number_format = fmt
                cells.append(cell)
            else:
                cells.append(value)
        ws.append(cells)

    return _save_and_stream(wb, filename)


# ─── CSV + ZIP helpers ──────────────────────────────────────────────────

def _csv_cell(v):
    """Coerce pandas sentinels to empty string for CSV (where '' = NULL)."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().isoformat(sep=" ")
    return v


def _stream_csv_file(tmp_path: str, filename: str, ext: str, media_type: str) -> StreamingResponse:
    """Stream a temp file back, deleting it after the read finishes."""
    def _it():
        try:
            with open(tmp_path, "rb") as f:
                while True:
                    chunk = f.read(65536)
                    if not chunk:
                        break
                    yield chunk
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    safe_name = re.sub(r'[\r\n"\\]', '', filename)
    headers = {"content-disposition": f'attachment; filename="{safe_name}.{ext}"'}
    return StreamingResponse(_it(), media_type=media_type, headers=headers)


def _write_csv_to_temp(columns, row_iter) -> str:
    """Write rows to a temp CSV; return the path. Caller is responsible for cleanup
    (which `_stream_csv_file` does on stream completion)."""
    tmp = tempfile.NamedTemporaryFile(prefix="vesign_export_", suffix=".csv", delete=False, mode="w", newline="", encoding="utf-8")
    try:
        writer = _csv.writer(tmp, quoting=_csv.QUOTE_MINIMAL)
        writer.writerow(columns)
        for row in row_iter:
            writer.writerow([_csv_cell(v) for v in row])
        tmp.flush()
        return tmp.name
    finally:
        tmp.close()


def _write_zip_csv_to_temp(columns, row_iter, csv_basename: str) -> str:
    """Write rows to a temp CSV inside a ZIP; return the .zip path."""
    csv_path = _write_csv_to_temp(columns, row_iter)
    zip_tmp = tempfile.NamedTemporaryFile(prefix="vesign_export_", suffix=".zip", delete=False)
    zip_tmp.close()
    try:
        with zipfile.ZipFile(zip_tmp.name, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            zf.write(csv_path, arcname=f"{csv_basename}.csv")
    finally:
        try:
            os.unlink(csv_path)
        except OSError:
            pass
    return zip_tmp.name


def dataframe_to_csv_response(df: pd.DataFrame, filename: str) -> StreamingResponse:
    """CSV download from a DataFrame. `filename` excludes extension."""
    columns = list(df.columns)
    rows = df.itertuples(index=False, name=None)
    tmp_path = _write_csv_to_temp(columns, rows)
    return _stream_csv_file(tmp_path, filename, "csv", CSV_MEDIA_TYPE)


def dataframe_to_zip_response(df: pd.DataFrame, filename: str) -> StreamingResponse:
    """ZIP-of-CSV download from a DataFrame. `filename` excludes extension."""
    columns = list(df.columns)
    rows = df.itertuples(index=False, name=None)
    tmp_path = _write_zip_csv_to_temp(columns, rows, csv_basename=filename)
    return _stream_csv_file(tmp_path, filename, "zip", ZIP_MEDIA_TYPE)


def cursor_to_csv_response(conn, sql, params: dict, filename: str) -> StreamingResponse:
    """CSV stream from a SQL cursor. Bypasses pandas — same OOM-safe pattern as cursor_to_xlsx_response."""
    result = conn.execute(sql, params)
    columns = list(result.keys())
    tmp_path = _write_csv_to_temp(columns, result)
    return _stream_csv_file(tmp_path, filename, "csv", CSV_MEDIA_TYPE)


def cursor_to_zip_response(conn, sql, params: dict, filename: str) -> StreamingResponse:
    """ZIP-of-CSV stream from a SQL cursor."""
    result = conn.execute(sql, params)
    columns = list(result.keys())
    tmp_path = _write_zip_csv_to_temp(columns, result, csv_basename=filename)
    return _stream_csv_file(tmp_path, filename, "zip", ZIP_MEDIA_TYPE)


# ─── Format dispatchers ─────────────────────────────────────────────────

def normalize_format(fmt: str | None) -> str:
    """Coerce the ?format= query param to one of SUPPORTED_FORMATS, default xlsx."""
    f = (fmt or "xlsx").strip().lower()
    return f if f in SUPPORTED_FORMATS else "xlsx"


def dispatch_dataframe_response(
    fmt: str,
    df: pd.DataFrame,
    filename: str,
    sheet_name: str = "Sheet1",
    column_formats: dict | None = None,
) -> StreamingResponse:
    """Pick the right dataframe-based responder for `fmt` (xlsx/csv/zip)."""
    fmt = normalize_format(fmt)
    if fmt == "xlsx":
        return dataframe_to_xlsx_response(df, filename, sheet_name=sheet_name, column_formats=column_formats)
    if fmt == "csv":
        return dataframe_to_csv_response(df, filename)
    return dataframe_to_zip_response(df, filename)


def dispatch_cursor_response(
    fmt: str,
    conn,
    sql,
    params: dict,
    filename: str,
    sheet_name: str = "Sheet1",
    column_formats: dict | None = None,
    date_columns: tuple = (),
    auto_size: bool = False,
) -> StreamingResponse:
    """Pick the right cursor-based responder for `fmt` (xlsx/csv/zip).

    XLSX uses formats + date parsing; CSV/ZIP write raw values (Excel/sheets
    interpret YYYY-MM-DD strings as dates anyway, and number formatting is
    lost in CSV by definition).

    `auto_size` only applies to XLSX — CSV/ZIP have no column-width concept.
    """
    fmt = normalize_format(fmt)
    if fmt == "xlsx":
        return cursor_to_xlsx_response(
            conn, sql, params, filename,
            sheet_name=sheet_name, column_formats=column_formats, date_columns=date_columns,
            auto_size=auto_size,
        )
    if fmt == "csv":
        return cursor_to_csv_response(conn, sql, params, filename)
    return cursor_to_zip_response(conn, sql, params, filename)
