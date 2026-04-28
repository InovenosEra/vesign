"""Smoke tests for backend.exports. Run with: venv/bin/python scripts/smoke_exports.py"""
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import asyncio
import inspect
import io
import subprocess
import time
import urllib.error
import urllib.request
import pandas as pd
from openpyxl import load_workbook

from backend.exports import dataframe_to_xlsx_response


def _read_response_body(resp):
    """Drain a FastAPI response into bytes, regardless of sync/async iterator."""
    if hasattr(resp, "body_iterator"):
        it = resp.body_iterator
        if inspect.isasyncgen(it):
            async def _drain():
                return b"".join([chunk async for chunk in it])
            return asyncio.run(_drain())
        return b"".join(it)
    return resp.body


def test_basic_workbook():
    df = pd.DataFrame([
        {"ticker": "AAPL", "close": 220.0, "rsi": 55.1},
        {"ticker": "MSFT", "close": 410.0, "rsi": 48.0},
    ])
    resp = dataframe_to_xlsx_response(df, filename="signals_2026-04-27", sheet_name="signals")

    assert resp.media_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'attachment; filename="signals_2026-04-27.xlsx"' in resp.headers["content-disposition"]

    body = _read_response_body(resp)
    wb = load_workbook(io.BytesIO(body))
    assert wb.sheetnames == ["signals"]
    ws = wb["signals"]
    assert [c.value for c in ws[1]] == ["ticker", "close", "rsi"]
    assert ws.cell(row=2, column=1).value == "AAPL"
    assert ws.cell(row=3, column=2).value == 410.0
    print("test_basic_workbook PASS")


def test_empty_dataframe_still_emits_header():
    df = pd.DataFrame(columns=["ticker", "close"])
    resp = dataframe_to_xlsx_response(df, filename="empty", sheet_name="signals")
    body = _read_response_body(resp)
    wb = load_workbook(io.BytesIO(body))
    ws = wb["signals"]
    assert [c.value for c in ws[1]] == ["ticker", "close"]
    assert ws.max_row == 1   # header only
    print("test_empty_dataframe_still_emits_header PASS")


def test_pandas_sentinels_serialize_cleanly():
    """NaN, NaT, pd.NA and Timestamp must coerce — they appear in every SQL-backed export."""
    df = pd.DataFrame([
        {"ticker": "AAPL", "close": 220.0, "buy_date": pd.Timestamp("2026-04-15")},
        {"ticker": "MSFT", "close": float("nan"), "buy_date": pd.NaT},
    ])
    resp = dataframe_to_xlsx_response(df, filename="sentinels", sheet_name="s")
    body = _read_response_body(resp)
    wb = load_workbook(io.BytesIO(body))
    ws = wb["s"]
    assert ws.cell(row=2, column=2).value == 220.0
    assert ws.cell(row=3, column=2).value is None   # NaN → None
    assert ws.cell(row=3, column=3).value is None   # NaT → None
    print("test_pandas_sentinels_serialize_cleanly PASS")


def _start_uvicorn():
    """Boot uvicorn in the background for HTTP smoke tests."""
    p = subprocess.Popen(
        ["venv/bin/uvicorn", "backend.main:app", "--port", "8765"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    # Wait for /api/health to respond
    for _ in range(60):
        try:
            with urllib.request.urlopen("http://127.0.0.1:8765/api/health", timeout=1) as r:
                if r.status == 200:
                    return p
        except Exception:
            time.sleep(0.5)
    p.terminate()
    raise RuntimeError("uvicorn did not start in 30s")


def _http_get(path: str, token: str | None = None) -> tuple[int, bytes, dict]:
    req = urllib.request.Request("http://127.0.0.1:8765" + path)
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return r.status, r.read(), dict(r.headers)
    except urllib.error.HTTPError as e:
        return e.code, e.read(), dict(e.headers)


def test_signals_export_requires_auth():
    status, body, _ = _http_get("/api/signals/export.xlsx")
    assert status == 401, f"expected 401 without token, got {status}: {body[:200]}"
    print("test_signals_export_requires_auth PASS")


def test_trades_export_requires_auth():
    status, body, _ = _http_get("/api/trades/export.xlsx")
    assert status == 401, f"expected 401, got {status}: {body[:200]}"
    print("test_trades_export_requires_auth PASS")


def test_open_trades_export_requires_auth():
    status, body, _ = _http_get("/api/trades/open/export.xlsx")
    assert status == 401, f"expected 401, got {status}: {body[:200]}"
    print("test_open_trades_export_requires_auth PASS")


def test_watchlist_export_requires_auth():
    status, body, _ = _http_get("/api/watchlists/1/export.xlsx")
    assert status == 401, f"expected 401, got {status}: {body[:200]}"
    print("test_watchlist_export_requires_auth PASS")


if __name__ == "__main__":
    test_basic_workbook()
    test_pandas_sentinels_serialize_cleanly()
    test_empty_dataframe_still_emits_header()
    print("\nAll exports.py smoke tests passed.")

    proc = _start_uvicorn()
    try:
        test_signals_export_requires_auth()
        test_trades_export_requires_auth()
        test_open_trades_export_requires_auth()
        test_watchlist_export_requires_auth()
    finally:
        proc.terminate()
        proc.wait(timeout=5)
